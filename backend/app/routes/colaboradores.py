from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
import re
import unicodedata

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from openpyxl import load_workbook
from pypdf import PdfReader
from sqlalchemy import or_
from sqlalchemy.orm import Session

from app.auth import exigir_perfis, obter_usuario_atual
from app.dependencies import get_db
from app.models import (
    Advertencia,
    AtestadoMedico,
    Colaborador,
    Falta,
    Suspensao,
)
from app.schemas import (
    ColaboradorCreate,
    ColaboradorDetalheResponse,
    ColaboradorOpcaoResponse,
    ColaboradorResponse,
    ColaboradorUpdate,
    ColaboradoresPaginadosResponse,
)

router = APIRouter(
    prefix="/colaboradores",
    tags=["Colaboradores"],
    dependencies=[Depends(obter_usuario_atual)]
)


def validar_identificadores_unicos(
    db: Session,
    cpf: str | None = None,
    matricula: str | None = None,
    colaborador_id: int | None = None,
):
    if cpf:
        query = db.query(Colaborador).filter(Colaborador.cpf == cpf)

        if colaborador_id is not None:
            query = query.filter(Colaborador.id != colaborador_id)

        if query.first():
            raise HTTPException(status_code=400, detail="CPF já cadastrado")

    if matricula:
        query = db.query(Colaborador).filter(Colaborador.matricula == matricula)

        if colaborador_id is not None:
            query = query.filter(Colaborador.id != colaborador_id)

        if query.first():
            raise HTTPException(status_code=400, detail="Matrícula já cadastrada")


def gerar_matricula_por_id(colaborador_id: int):
    return f"{colaborador_id:06d}"


def aplicar_regras_desligamento(campos: dict):
    if campos.get("data_desligamento"):
        campos["ativo"] = False


def normalizar_texto(valor):
    texto = str(valor or "").strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(char for char in texto if not unicodedata.combining(char))
    texto = re.sub(r"[^a-z0-9]+", "_", texto)
    return texto.strip("_")


def limpar_texto(valor):
    if valor is None:
        return None

    texto = str(valor).strip()

    if normalizar_texto(texto) in {
        "nan",
        "none",
        "null",
        "nulo",
        "sem_email",
        "sem_telefone",
        "nao_informado",
        "nao_informada",
    }:
        return None

    if re.fullmatch(r"[-_.\s]+", texto):
        return None

    return texto or None


def normalizar_email_importacao(valor):
    texto = limpar_texto(valor)

    if not texto:
        return None

    texto = texto.lower()

    if not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", texto):
        return None

    return texto


def normalizar_telefone_importacao(valor):
    texto = limpar_texto(valor)

    if not texto:
        return None

    digitos = re.sub(r"\D", "", texto)

    return digitos or None


def normalizar_cpf_importacao(valor):
    texto = limpar_texto(valor)

    if not texto:
        return None

    digitos = re.sub(r"\D", "", texto)

    if len(digitos) == 10:
        digitos = digitos.zfill(11)

    if len(digitos) != 11:
        return digitos

    return f"{digitos[:3]}.{digitos[3:6]}.{digitos[6:9]}-{digitos[9:]}"


def normalizar_rg_importacao(valor):
    texto = limpar_texto(valor)

    if not texto:
        return None

    return re.sub(r"\D", "", texto) or texto


def normalizar_empresa(valor):
    texto = normalizar_texto(valor)

    if texto in {"c_m", "cm"}:
        return "C&M"

    if texto == "frontline":
        return "Frontline"

    return None


def normalizar_data(valor):
    if valor in (None, ""):
        return None

    if isinstance(valor, datetime):
        return valor.date()

    if isinstance(valor, date):
        return valor

    texto = str(valor).strip()
    digitos = re.sub(r"\D", "", texto)

    if len(digitos) == 8:
        try:
            return datetime.strptime(digitos, "%d%m%Y").date()
        except ValueError:
            pass

    for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            pass

    raise ValueError(f"Data inválida: {texto}")


def adicionar_meses(data_base: date, meses: int):
    mes_total = data_base.month - 1 + meses
    ano = data_base.year + mes_total // 12
    mes = mes_total % 12 + 1
    ultimo_dia_mes = (
        date(ano + 1, 1, 1)
        if mes == 12
        else date(ano, mes + 1, 1)
    ).toordinal() - date(ano, mes, 1).toordinal()

    return date(ano, mes, min(data_base.day, ultimo_dia_mes))


def extrair_texto_pdf(conteudo: bytes):
    try:
        leitor = PdfReader(BytesIO(conteudo))
        return "\n".join(pagina.extract_text() or "" for pagina in leitor.pages)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="PDF inválido") from exc


def detectar_empresa_relatorio_ferias(texto: str):
    texto_normalizado = normalizar_texto(texto)

    if "frontline_servicos" in texto_normalizado:
        return "Frontline"

    if (
        "c_e_m_distribuidora" in texto_normalizado
        or "c_m_distribuidora" in texto_normalizado
    ):
        return "C&M"

    return None


def extrair_registros_ferias_pdf(texto: str):
    linhas = [linha.strip() for linha in texto.splitlines() if linha.strip()]
    registros = []

    for indice, linha in enumerate(linhas):
        if not re.fullmatch(r"\d{6}", linha):
            continue

        if indice + 1 >= len(linhas):
            continue

        nome = limpar_texto(linhas[indice + 1])
        if not nome:
            continue

        periodo_indice = None
        for candidato in range(indice + 2, min(indice + 12, len(linhas))):
            if re.fullmatch(
                r"\d{2}/\d{2}/\d{4}\s+a\s+\d{2}/\d{2}/\d{4}",
                linhas[candidato],
            ):
                periodo_indice = candidato
                break

        if periodo_indice is None:
            continue

        try:
            inicio_periodo_texto, fim_periodo_texto = re.split(
                r"\s+a\s+",
                linhas[periodo_indice],
                maxsplit=1,
            )
            inicio_periodo = normalizar_data(inicio_periodo_texto)
            fim_periodo = normalizar_data(fim_periodo_texto)
        except ValueError:
            continue

        registros.append({
            "codigo": linha,
            "nome": nome,
            "data_inicio_periodo_aquisitivo": inicio_periodo,
            "data_fim_periodo_aquisitivo": fim_periodo,
            "data_limite_ferias": adicionar_meses(fim_periodo, 9),
        })

    registros_por_nome = {}

    for registro in registros:
        chave = normalizar_texto(registro["nome"])
        atual = registros_por_nome.get(chave)

        if (
            not atual
            or registro["data_fim_periodo_aquisitivo"] < atual["data_fim_periodo_aquisitivo"]
        ):
            registros_por_nome[chave] = registro

    return list(registros_por_nome.values())


def normalizar_tipo_contrato(valor):
    texto = normalizar_texto(valor)

    contratos = {
        "clt": "CLT",
        "pj": "PJ",
        "temporario": "Temporario",
        "estagio": "Estagio",
        "terceirizado": "Terceirizado",
    }

    return contratos.get(texto, limpar_texto(valor))


def normalizar_tipo_bonificacao(valor):
    texto = normalizar_texto(valor)

    tipos = {
        "fixa": "Fixa",
        "fixo": "Fixa",
        "mensal": "Variavel",
        "variavel": "Variavel",
        "variavel_mensal": "Variavel",
    }

    return tipos.get(texto, limpar_texto(valor))


def normalizar_salario(valor):
    if valor in (None, ""):
        return None

    if isinstance(valor, (int, float, Decimal)):
        return Decimal(str(valor)).quantize(Decimal("0.01"))

    texto = str(valor).strip()
    texto = re.sub(r"[^\d,.-]", "", texto)

    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")

    try:
        return Decimal(texto).quantize(Decimal("0.01"))
    except InvalidOperation as exc:
        raise ValueError(f"Salário inválido: {valor}") from exc


COLUNAS_IMPORTACAO = {
    "empresa": "empresa",
    "nome": "nome",
    "matricula": "matricula",
    "cargo": "cargo",
    "salario": "salario",
    "tipo_de_bonificacao": "tipo_bonificacao",
    "tipo_bonificacao": "tipo_bonificacao",
    "bonificacao": "bonificacao",
    "setor": "setor",
    "tipo_de_contrato": "tipo_contrato",
    "cpf": "cpf",
    "rg": "rg",
    "data_de_nascimento": "data_nascimento",
    "data_de_admissao": "data_admissao",
    "data_do_aso": "data_aso",
    "inicio_do_periodo_aquisitivo": "data_inicio_periodo_aquisitivo",
    "inicio_periodo_aquisitivo": "data_inicio_periodo_aquisitivo",
    "data_inicio_periodo_aquisitivo": "data_inicio_periodo_aquisitivo",
    "fim_do_periodo_aquisitivo": "data_fim_periodo_aquisitivo",
    "fim_periodo_aquisitivo": "data_fim_periodo_aquisitivo",
    "data_fim_periodo_aquisitivo": "data_fim_periodo_aquisitivo",
    "data_limite_de_ferias": "data_limite_ferias",
    "data_limite_ferias": "data_limite_ferias",
    "e_mail": "email",
    "email": "email",
    "telefone": "telefone",
    "telefone_de_emergencia": "telefone_emergencia",
    "endereco": "endereco",
}


def montar_campos_importacao(headers, valores):
    campos = {}

    for indice, header in enumerate(headers):
        campo = COLUNAS_IMPORTACAO.get(normalizar_texto(header))

        if not campo:
            continue

        valor = valores[indice] if indice < len(valores) else None

        if campo == "matricula":
            continue

        if campo == "empresa":
            campos[campo] = normalizar_empresa(valor)
        elif campo == "cpf":
            campos[campo] = normalizar_cpf_importacao(valor)
        elif campo == "rg":
            campos[campo] = normalizar_rg_importacao(valor)
        elif campo == "tipo_contrato":
            campos[campo] = normalizar_tipo_contrato(valor)
        elif campo == "tipo_bonificacao":
            campos[campo] = normalizar_tipo_bonificacao(valor)
        elif campo == "salario":
            campos[campo] = normalizar_salario(valor)
        elif campo == "bonificacao":
            campos[campo] = normalizar_salario(valor)
        elif campo.startswith("data_"):
            campos[campo] = normalizar_data(valor)
        elif campo == "email":
            campos[campo] = normalizar_email_importacao(valor)
        elif campo in {"telefone", "telefone_emergencia"}:
            campos[campo] = normalizar_telefone_importacao(valor)
        else:
            campos[campo] = limpar_texto(valor)

    return campos


@router.post("/", response_model=ColaboradorResponse)
def criar_colaborador(
    colaborador: ColaboradorCreate,
    _usuario=Depends(exigir_perfis("admin", "rh")),
    db: Session = Depends(get_db)
):
    validar_identificadores_unicos(
        db,
        cpf=colaborador.cpf,
    )

    campos = colaborador.model_dump()
    campos.pop("matricula", None)
    aplicar_regras_desligamento(campos)

    novo_colaborador = Colaborador(
        **campos
    )

    db.add(novo_colaborador)
    db.flush()
    novo_colaborador.matricula = gerar_matricula_por_id(novo_colaborador.id)
    db.commit()
    db.refresh(novo_colaborador)

    return novo_colaborador


@router.post("/importar")
async def importar_colaboradores(
    arquivo: UploadFile = File(...),
    _usuario=Depends(exigir_perfis("admin", "rh")),
    db: Session = Depends(get_db),
):
    if not arquivo.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Envie um arquivo .xlsx")

    conteudo = await arquivo.read()

    try:
        workbook = load_workbook(BytesIO(conteudo), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Planilha inválida") from exc

    sheet = workbook[workbook.sheetnames[0]]
    linhas = list(sheet.iter_rows(values_only=True))

    if not linhas:
        raise HTTPException(status_code=400, detail="Planilha vazia")

    headers = linhas[0]
    importados = 0
    ignorados = 0
    linhas_ignoradas = []
    erros = []
    cpfs_planilha = set()

    for numero_linha, valores in enumerate(linhas[1:], start=2):
        try:
            campos = montar_campos_importacao(headers, valores)

            if not campos.get("nome"):
                ignorados += 1
                linhas_ignoradas.append({
                    "linha": numero_linha,
                    "motivo": "Nome não informado",
                })
                continue

            campos["empresa"] = campos.get("empresa") or "C&M"

            cpf = campos.get("cpf")
            if cpf:
                if cpf in cpfs_planilha:
                    raise ValueError("CPF duplicado na planilha")

                cpfs_planilha.add(cpf)

            colaborador_validado = ColaboradorCreate(**campos)
            validar_identificadores_unicos(
                db,
                cpf=colaborador_validado.cpf,
            )

            dados = colaborador_validado.model_dump()
            dados.pop("matricula", None)
            aplicar_regras_desligamento(dados)

            novo_colaborador = Colaborador(**dados)
            db.add(novo_colaborador)
            db.flush()
            novo_colaborador.matricula = gerar_matricula_por_id(
                novo_colaborador.id
            )
            db.commit()
            importados += 1
        except Exception as exc:
            db.rollback()
            erros.append({
                "linha": numero_linha,
                "erro": str(exc),
            })

    return {
        "importados": importados,
        "ignorados": ignorados,
        "linhas_ignoradas": linhas_ignoradas,
        "erros": erros,
    }


@router.post("/importar-ferias-pdf")
async def importar_limite_ferias_pdf(
    arquivo: UploadFile = File(...),
    _usuario=Depends(exigir_perfis("admin", "rh")),
    db: Session = Depends(get_db),
):
    if not arquivo.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Envie um arquivo .pdf")

    conteudo = await arquivo.read()
    texto = extrair_texto_pdf(conteudo)
    registros = extrair_registros_ferias_pdf(texto)

    if not registros:
        raise HTTPException(
            status_code=400,
            detail="Nenhum funcionário com período aquisitivo foi encontrado no PDF",
        )

    empresa = detectar_empresa_relatorio_ferias(texto)
    colaboradores = db.query(Colaborador).all()
    colaboradores_por_nome = {}

    for colaborador in colaboradores:
        colaboradores_por_nome.setdefault(
            normalizar_texto(colaborador.nome),
            [],
        ).append(colaborador)

    importados = 0
    ignorados = 0
    erros = []

    for indice, registro in enumerate(registros, start=1):
        chave_nome = normalizar_texto(registro["nome"])
        candidatos = colaboradores_por_nome.get(chave_nome, [])

        if empresa:
            candidatos_empresa = [
                colaborador
                for colaborador in candidatos
                if colaborador.empresa == empresa
            ]
            candidatos = candidatos_empresa or candidatos

        if not candidatos:
            ignorados += 1
            erros.append({
                "linha": indice,
                "erro": f"Colaborador não encontrado: {registro['nome']}",
            })
            continue

        candidatos_ativos = [
            colaborador
            for colaborador in candidatos
            if colaborador.ativo
        ]

        if len(candidatos) > 1 and len(candidatos_ativos) == 1:
            colaborador = candidatos_ativos[0]
        elif len(candidatos) == 1:
            colaborador = candidatos[0]
        else:
            ignorados += 1
            erros.append({
                "linha": indice,
                "erro": f"Mais de um colaborador encontrado: {registro['nome']}",
            })
            continue

        colaborador.data_inicio_periodo_aquisitivo = registro[
            "data_inicio_periodo_aquisitivo"
        ]
        colaborador.data_fim_periodo_aquisitivo = registro[
            "data_fim_periodo_aquisitivo"
        ]
        colaborador.data_limite_ferias = registro["data_limite_ferias"]
        importados += 1

    db.commit()

    return {
        "importados": importados,
        "ignorados": ignorados,
        "erros": erros,
        "empresa": empresa,
    }


@router.delete("/importar-ferias-pdf")
def limpar_importacao_ferias_pdf(
    _usuario=Depends(exigir_perfis("admin", "rh")),
    db: Session = Depends(get_db),
):
    query = db.query(Colaborador).filter(
        or_(
            Colaborador.data_inicio_periodo_aquisitivo.isnot(None),
            Colaborador.data_fim_periodo_aquisitivo.isnot(None),
            Colaborador.data_limite_ferias.isnot(None),
        )
    )
    total = query.count()

    query.update(
        {
            Colaborador.data_inicio_periodo_aquisitivo: None,
            Colaborador.data_fim_periodo_aquisitivo: None,
            Colaborador.data_limite_ferias: None,
        },
        synchronize_session=False,
    )
    db.commit()

    return {"limpos": total}


@router.get("/", response_model=list[ColaboradorResponse])
def listar_colaboradores(
    db: Session = Depends(get_db)
):
    return db.query(Colaborador).all()


@router.get("/opcoes", response_model=list[ColaboradorOpcaoResponse])
def listar_opcoes_colaboradores(
    ativo: bool | None = None,
    db: Session = Depends(get_db)
):
    query = db.query(Colaborador)

    if ativo is not None:
        query = query.filter(Colaborador.ativo == ativo)

    return query.order_by(Colaborador.nome).all()


@router.get("/busca", response_model=ColaboradoresPaginadosResponse)
def buscar_colaboradores(
    q: str | None = None,
    status: str = "todos",
    skip: int = 0,
    limit: int = 10,
    db: Session = Depends(get_db)
):
    skip = max(skip, 0)
    limit = min(max(limit, 1), 100)

    query = db.query(Colaborador)

    if q:
        termo = f"%{q.strip()}%"
        query = query.filter(
            or_(
                Colaborador.nome.ilike(termo),
                Colaborador.empresa.ilike(termo),
                Colaborador.matricula.ilike(termo),
                Colaborador.cargo.ilike(termo),
                Colaborador.setor.ilike(termo),
                Colaborador.cpf.ilike(termo),
                Colaborador.telefone.ilike(termo),
            )
        )

    if status == "ativos":
        query = query.filter(Colaborador.ativo == True)
    elif status == "inativos":
        query = query.filter(Colaborador.ativo == False)

    total = query.count()
    items = (
        query
        .order_by(Colaborador.nome)
        .offset(skip)
        .limit(limit)
        .all()
    )

    return {
        "items": items,
        "total": total,
        "skip": skip,
        "limit": limit,
    }


@router.put("/{colaborador_id}", response_model=ColaboradorResponse)
def atualizar_colaborador(
    colaborador_id: int,
    dados: ColaboradorUpdate,
    _usuario=Depends(exigir_perfis("admin", "rh")),
    db: Session = Depends(get_db)
):
    colaborador = db.query(Colaborador).filter(
        Colaborador.id == colaborador_id
    ).first()

    if not colaborador:
        raise HTTPException(status_code=404, detail="Colaborador não encontrado")

    campos = dados.model_dump(exclude_unset=True)
    campos.pop("matricula", None)
    aplicar_regras_desligamento(campos)

    validar_identificadores_unicos(
        db,
        cpf=campos.get("cpf"),
        colaborador_id=colaborador_id,
    )

    for campo, valor in campos.items():
        setattr(colaborador, campo, valor)

    db.commit()
    db.refresh(colaborador)

    return colaborador


@router.patch("/{colaborador_id}/inativar", response_model=ColaboradorResponse)
def inativar_colaborador(
    colaborador_id: int,
    _usuario=Depends(exigir_perfis("admin", "rh")),
    db: Session = Depends(get_db)
):
    colaborador = db.query(Colaborador).filter(
        Colaborador.id == colaborador_id
    ).first()

    if not colaborador:
        raise HTTPException(status_code=404, detail="Colaborador não encontrado")

    colaborador.ativo = False

    db.commit()
    db.refresh(colaborador)

    return colaborador


@router.patch("/{colaborador_id}/ativar", response_model=ColaboradorResponse)
def ativar_colaborador(
    colaborador_id: int,
    _usuario=Depends(exigir_perfis("admin", "rh")),
    db: Session = Depends(get_db)
):
    colaborador = db.query(Colaborador).filter(
        Colaborador.id == colaborador_id
    ).first()

    if not colaborador:
        raise HTTPException(status_code=404, detail="Colaborador não encontrado")

    colaborador.ativo = True

    db.commit()
    db.refresh(colaborador)

    return colaborador


@router.get("/{colaborador_id}", response_model=ColaboradorDetalheResponse)
def detalhar_colaborador(
    colaborador_id: int,
    db: Session = Depends(get_db)
):
    colaborador = db.query(Colaborador).filter(
        Colaborador.id == colaborador_id
    ).first()

    if not colaborador:
        raise HTTPException(
            status_code=404,
            detail="Colaborador não encontrado"
        )

    faltas = db.query(Falta).filter(
        Falta.colaborador_id == colaborador_id,
        Falta.removido_em.is_(None),
    ).all()

    advertencias = db.query(Advertencia).filter(
        Advertencia.colaborador_id == colaborador_id,
        Advertencia.removido_em.is_(None),
    ).all()

    suspensoes = db.query(Suspensao).filter(
        Suspensao.colaborador_id == colaborador_id,
        Suspensao.removido_em.is_(None),
    ).all()

    atestados = db.query(AtestadoMedico).filter(
        AtestadoMedico.colaborador_id == colaborador_id,
        AtestadoMedico.removido_em.is_(None),
    ).all()

    return {
        "colaborador": ColaboradorResponse.model_validate(colaborador),
        "faltas": faltas,
        "advertencias": advertencias,
        "suspensoes": suspensoes,
        "atestados": atestados
    }
